import os
from openpyxl import load_workbook, Workbook

# Thư mục chứa các file Excel cần gộp
folder_path = "C:/dangnv20/KPI_KQI_Toi/SYSLOG"

# Tạo Workbook mới để lưu kết quả gộp
combined_workbook = Workbook()
combined_sheet = combined_workbook.active
combined_sheet.title = "Combined"

# Duyệt qua tất cả các file Excel trong thư mục
for filename in os.listdir(folder_path):
    if filename.endswith(".xlsx"):
        # Đọc file Excel
        file_path = os.path.join(folder_path, filename)
        workbook = load_workbook(file_path)
        sheet = workbook.active  # Lấy sheet đầu tiên trong file

        # Sao chép dữ liệu từ file hiện tại vào file tổng hợp
        for row in sheet.iter_rows(values_only=True):
            combined_sheet.append(row)

# Lưu file Excel sau khi gộp
combined_workbook.save("C:/dangnv20/KPI_KQI_Toi/SYSLOG/syslog_AGG_BLU_HUG_LAN_TVH.xlsx")

print("Đã gộp file Excel thành công!")
