import openpyxl
import shutil
import os

def copy_and_split_excel_by_condition(file_path, sheet_name, condition_column, condition_values):
    # Kiểm tra file có tồn tại hay không
    if not os.path.isfile(file_path):
        print(f"File '{file_path}' không tồn tại.")
        return
    
    print(f"Đang mở file: {file_path}")
    
    # Mở file Excel gốc
    wb = openpyxl.load_workbook(file_path)
    print(f"Đã mở file: {file_path}")
    
    # Kiểm tra xem sheet có tồn tại hay không
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file '{file_path}'.")
        return
    
    sheet = wb[sheet_name]
    print(f"Đã mở sheet: {sheet_name}")

    # Đọc dữ liệu từ sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Tách dữ liệu theo điều kiện
    header = data[:1]  # Giữ nguyên header
    rows = data[1:]
    grouped_rows = {value: [] for value in condition_values}
    
    for row in rows:
        if row[condition_column] in condition_values:
            grouped_rows[row[condition_column]].append(row)
    
    # Hàm để tạo file mới từ dữ liệu
    def create_new_file(rows, suffix):
        new_file_path = os.path.join(os.path.dirname(file_path), f'{suffix}.xlsx')
        shutil.copyfile(file_path, new_file_path)
        new_wb = openpyxl.load_workbook(new_file_path)
        new_sheet = new_wb[sheet_name]
        new_sheet.delete_rows(1, new_sheet.max_row)
        for j, row in enumerate(rows, start=1):
            for k, value in enumerate(row, start=1):
                new_sheet.cell(row=j, column=k, value=value)
        new_wb.save(new_file_path)
        print(f"Đã lưu: {new_file_path}")

    # Tạo file cho các dòng thỏa mãn từng điều kiện
    for condition_value, rows in grouped_rows.items():
        if rows:
            create_new_file(header + rows, condition_value)

# Đường dẫn tới file Excel gốc và tên sheet cần tách
file_path = 'C:/Users/vdang/OneDrive/Documents/WebView/tool_tach_file_theo_dkien.xlsx'  # Đường dẫn tuyệt đối
sheet_name = 'Sheet1'
condition_column = 3  # Cột điều kiện (index bắt đầu từ 0)
condition_values = ['CR3']  # Các giá trị điều kiện

copy_and_split_excel_by_condition(file_path, sheet_name, condition_column, condition_values)
