import openpyxl
import shutil
import os

def copy_and_split_excel(file_path, sheet_name, chunk_size):
    # Kiểm tra file có tồn tại hay không
    if not os.path.isfile(file_path):
        print(f"File '{file_path}' không tồn tại.")
        return
    
    print(f"Đang mở file: {file_path}")
    
    # Mở file Excel gốc
    wb = openpyxl.load_workbook(file_path)
    print(f"Đã mở file: {file_path}")
    
    # Kiểm tra xem sheet có tồn tại hay không
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file '{file_path}'.")
        return
    
    sheet = wb[sheet_name]
    print(f"Đã mở sheet: {sheet_name}")

    # Đọc dữ liệu từ sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Tách dữ liệu thành các file mới
    header = data[:6]  # Giữ nguyên 6 hàng đầu
    rows = data[6:]
    num_chunks = len(rows) // chunk_size + (1 if len(rows) % chunk_size != 0 else 0)

    for i in range(num_chunks):
        start_row = i * chunk_size
        end_row = (i + 1) * chunk_size
        chunk = rows[start_row:end_row]
        
        # Sao chép file gốc thành file mới
        new_file_path = os.path.join(os.path.dirname(file_path), f'MOP_reboot_{i+1}.xlsx')
        shutil.copyfile(file_path, new_file_path)
        print(f"Đã sao chép file: {new_file_path}")
        
        # Mở file mới và ghi dữ liệu đã tách vào
        new_wb = openpyxl.load_workbook(new_file_path)
        new_sheet = new_wb[sheet_name]
        
        # Xóa toàn bộ dữ liệu cũ từ dòng thứ 7
        new_sheet.delete_rows(7, new_sheet.max_row)
        
        # Ghi dữ liệu mới vào sheet bắt đầu từ dòng thứ 7
        for j, row in enumerate(chunk, start=7):
            for k, value in enumerate(row, start=1):
                new_sheet.cell(row=j, column=k, value=value)
        
        # Lưu file mới
        new_wb.save(new_file_path)
        print(f"Đã lưu: {new_file_path}")

# Đường dẫn tới file Excel gốc và tên sheet cần tách
file_path = 'C:/Users/vdang/OneDrive/Documents/Dangnv20/tool_tach_file/data.xlsx'  # Đường dẫn tuyệt đối
sheet_name = 'GPON_OLT-ZTE-C320_mini_OLT_ZTE'
chunk_size = 100  # Số hàng trong mỗi file nhỏ

copy_and_split_excel(file_path, sheet_name, chunk_size)
