import openpyxl
from openpyxl import load_workbook, Workbook
import os

def split_sheet_and_overwrite(file_path, sheet_name, rows_per_file, output_files):
    # Mở file Excel gốc
    wb = load_workbook(file_path)
    
    # Kiểm tra xem sheet có tồn tại không
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file.")
        return
    
    ws = wb[sheet_name]
    
    # Đọc dữ liệu từ dòng thứ 7 trở đi
    data = list(ws.iter_rows(min_row=13, values_only=True))
    
    # Tổng số dòng dữ liệu bắt đầu từ dòng thứ 7
    total_rows = len(data)
    
    # Đọc danh sách các file sẵn có
    for i, output_file in enumerate(output_files):
        # Xác định phạm vi dòng cần ghi cho mỗi file
        start_row = i * rows_per_file
        end_row = min(start_row + rows_per_file, total_rows)
        
        if start_row >= total_rows:
            print(f"Không có đủ dữ liệu để ghi vào file {output_file}.")
            continue
        
        # Tạo workbook mới hoặc mở file sẵn có
        if os.path.exists(output_file):
            new_wb = load_workbook(output_file)
        else:
            new_wb = Workbook()
        
        # Tạo sheet mới hoặc chọn sheet đã có
        if sheet_name in new_wb.sheetnames:
            new_ws = new_wb[sheet_name]
        else:
            new_ws = new_wb.create_sheet(title=sheet_name)
        
        # Nếu sheet đã có dữ liệu, xóa dữ liệu từ dòng thứ 7
        if new_ws.max_row >= 13:
            new_ws.delete_rows(13, 22 - 12)
        
        # Ghi dữ liệu vào dòng thứ 7 của file tách
        for row in data[start_row:end_row]:
            new_ws.append(row)
        
        # Lưu file mới
        new_wb.save(output_file)
        print(f"Đã ghi đè vào sheet '{sheet_name}' của file {output_file}")

if __name__ == "__main__":
    # Thay đổi đường dẫn file Excel gốc, tên sheet, số dòng mỗi file và danh sách file sẵn có
    file_path = 'D:/2024-2025/NangCapOLT_Dasan/CHECKLIST_new.xlsx'
    sheet_name = 'Checklist'  # Tên sheet cần tách
    rows_per_file = 10  # Số dòng mỗi file
    output_files = [
        "D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_1.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_2.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_3.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_4.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_5.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_6.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_7.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_8.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_9.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_10.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_11.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_12.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_13.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_14.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_15.xlsx",
"D:/2024-2025/NangCapOLT_Dasan/Checklist_CR_16.xlsx"
    ]

    # Chạy chức năng tách sheet và ghi đè vào các file sẵn có
    split_sheet_and_overwrite(file_path, sheet_name, rows_per_file, output_files)

    

    

    

   

        
        

