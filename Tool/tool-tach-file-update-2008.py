import openpyxl
import shutil
import os
from copy import copy  # Thêm dòng này để import hàm copy
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Alignment, Font

def copy_and_split_excel_by_condition(file_path, sheet_name, condition_column):
    if not os.path.isfile(file_path):
        print(f"File '{file_path}' không tồn tại.")
        return
    
    print(f"Đang mở file: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    print(f"Đã mở file: {file_path}")
    
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' không tồn tại trong file '{file_path}'.")
        return
    
    sheet = wb[sheet_name]
    print(f"Đã mở sheet: {sheet_name}")

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    header = data[:6]
    rows = data[6:]
    grouped_rows = {}
    
    for row in rows:
        condition_value = row[condition_column]
        if condition_value not in grouped_rows:
            grouped_rows[condition_value] = []
        grouped_rows[condition_value].append(row)
    
    def copy_cell_format(source_cell, target_cell):
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

    def create_new_file(rows, suffix, row_count):
        new_file_path = os.path.join(os.path.dirname(file_path), f'{suffix}.xlsx')
        shutil.copyfile(file_path, new_file_path)
        new_wb = openpyxl.load_workbook(new_file_path)
        new_sheet = new_wb[sheet_name]
        new_sheet.delete_rows(1, new_sheet.max_row)  # Xóa toàn bộ dữ liệu cũ từ dòng thứ 2
        
        for i, row in enumerate(rows, start=1):
            for j, value in enumerate(row, start=1):
                cell = new_sheet.cell(row=i, column=j, value=value)
                if i == 1:  # Copy định dạng của header
                    source_cell = sheet.cell(row=1, column=j)
                    copy_cell_format(source_cell, cell)
        
        new_wb.save(new_file_path)
        print(f"Đã lưu: {new_file_path} (Số dòng: {row_count})")

    for condition_value, rows in grouped_rows.items():
        if rows:
            row_count = len(rows)
            create_new_file(header + rows, condition_value, row_count)

# Đường dẫn tới file Excel gốc và tên sheet cần tách
file_path = 'D:/2024-2025/Xoa_uplink_OLT/template/C300/template_C300.xlsx'  # Đường dẫn tuyệt đối
sheet_name = 'SRT-Cisco-ASR920-24SZ-IM'
condition_column = 3  # Cột điều kiện (index bắt đầu từ 0)

copy_and_split_excel_by_condition(file_path, sheet_name, condition_column)
